#FUNCIONALIDADES DO CÓDIGO:
#INTERFACE DE SELEÇÃO AUTOMÁTICA (ABRE JANELAS AUTOMÁTICAS PARA A SELEÇÃO DA PLANILHA E DA PASTA DE SALVAMENTO)
#CRIAÇÃO DE BACKUP AUTOMÁTICO (SALVA UMA CÓPIA DE SEGURANÇA AUTOMÁTICA DO EXCEL ORIGINAL NA PASTA ESCOLHIDA, E EXIBE NO CONSOLE O CAMINHO COMPLETO DO BACKUP CRIADO)
#DETECÇÃO AUTOMÁTICA DO TIPO DE PLANILHA (MULTI-CÉLULA OU LINHA-A-LINHA)
#INDEXAÇÃO INTELIGENTE DE DOCUMENTOS (CRIA UM ÍNDICE INTERNO COM A VERSÃO NORMALIZADA DE CADA DOCUMENTO, TAMBÉM É UM PONTEIRO EXATO PARA LOCALIZA-LO, QUE PERMITE QUE AS BUSCAS SEJAM MAIS RÁPIDAS)
#NORMALIZAÇÃO DOS DADOS (PADRONIZA OS DOCUMENTOS PARA EVITAR FALHAS, REMOVE: PREFIXOS COMO, CTE, CTE RODOVIÁRIO, ETC. REMOVE LETRAS, HIFENS, ESPAÇOES E CARACTERES ESPECIAIS, SUFIXO COMO: /1.
#MODO INTERATIVO EM TMEPO REAL (O PROGRAMA FUNCIONA EM MODO CONVERSACIONAL COM O USUÁRIO)
#MARCAÇÃO PRECISA NO EXCEL (QUANDO LOCALIZADO PREENCHE COM: [✅ 1642742/1])
#CONTROLE DE DUPLICIDADES (SE O USUÁRIO DIGITAR O MESMO NÚMERO MAIS DE UMA VEZ: O PROGRAMA RECONHECE AUTOMATICAMENTE E MOSTRA)
#LOG DETALHADO DE CONFERÊNCIA
#RELATÓRIO FINAL AUTOMÁTICO (AO DIGITAR "FIM", O PROGRAMA MOSTRA UM RESUMO COMPLETO)
#SALVAMENTO INTELIGENTE (O RESULTADO DA CONFERÊNCIA É SALVO AUTOMATICAMENTE
#ESTRUTURA MODULA E EXTENSÍVEL (O CÓDIGO ESTÁ DIVIDO DE FORMA LIMPA)
#=======================================================================================================================

from openpyxl import load_workbook
from tkinter import Tk, filedialog
from pathlib import Path
import shutil
import re
from datetime import datetime
import sys


# ==========================================================
# 🔧 Funções utilitárias
# ==========================================================
def normalizar(texto):
    """Remove prefixos e caracteres desnecessários para comparação."""
    if not texto:
        return ""
    texto = str(texto).strip().upper()
    texto = re.sub(r"\bCT[-\s]*E\b.*?:", "", texto, flags=re.IGNORECASE)
    texto = re.sub(r"[A-ZÀ-Ú\s:\-]+", "", texto)
    texto = re.sub(r"[^0-9/]", "", texto)
    texto = re.sub(r"/1$", "", texto)
    return texto.strip().lower()


def detectar_tipo_planilha(aba):
    """Detecta se a planilha é multi-célula (vários documentos por célula) ou linha-a-linha."""
    for row in aba.iter_rows(min_row=1, max_row=min(30, aba.max_row), values_only=True):
        for valor in row:
            if isinstance(valor, str) and "," in valor:
                return "multi_celula"
    return "linha_a_linha"


def gerar_log(pasta_backup, arquivo, doc, status):
    """Gera log de conferência."""
    pasta_backup.mkdir(parents=True, exist_ok=True)
    log_path = pasta_backup / f"{arquivo.stem}_log.txt"
    with open(log_path, "a", encoding="utf-8") as log:
        log.write(f"[{datetime.now().strftime('%d/%m/%Y %H:%M:%S')}] - {doc} -> {status}\n")


# ==========================================================
# ⚙️ Função principal de conferência
# ==========================================================
def conferir_documentos_interativo(arquivo_excel, pasta_backup):
    try:
        arquivo = Path(arquivo_excel)
        if not arquivo.exists():
            print("❌ Arquivo não encontrado.")
            return

        # Criação do backup
        pasta_backup.mkdir(parents=True, exist_ok=True)
        backup = pasta_backup / f"{arquivo.stem}_backup{arquivo.suffix}"
        shutil.copy(arquivo, backup)
        print(f"🗂️ Backup criado em: {backup}")

        wb = load_workbook(arquivo)
        aba = wb.active

        tipo = detectar_tipo_planilha(aba)
        print(f"\n📊 Tipo de planilha detectado: {tipo.upper()}")

        # ==========================================================
        # 🧠 ETAPA 1: Indexação de documentos
        # ==========================================================
        print("🔍 Indexando documentos da planilha...")

        indice_docs = {}
        linha_inicial = 1

        for row in aba.iter_rows(min_row=1, max_row=aba.max_row):
            for celula in row:
                if not celula.value:
                    continue
                valor = str(celula.value)
                docs = [d.strip() for d in valor.split(",") if d.strip()]
                for doc in docs:
                    doc_norm = normalizar(doc)
                    if doc_norm:
                        indice_docs[doc_norm] = celula

        print(f"✅ Indexação concluída ({len(indice_docs)} documentos identificados).")

        # ==========================================================
        # 🟢 Modo interativo
        # ==========================================================
        documentos_encontrados = set()
        documentos_nao_encontrados = set()
        documentos_ja_conferidos = set()

        print("\n🟢 Modo interativo iniciado.")
        print("➡️ Digite o número do documento (ou 'fim' para encerrar).")

        while True:
            doc_input = input("\nDocumento: ").strip()
            if doc_input.lower() == "fim":
                break
            if not re.match(r"^\d+(\/\d+)?$", doc_input):
                print("⚠️ Entrada inválida. Digite apenas números (ex: 123456 ou 123456/1).")
                continue

            doc_norm = normalizar(doc_input)

            # Já conferido anteriormente
            if doc_norm in documentos_encontrados:
                print(f"⚠️ Documento {doc_input} já conferido anteriormente.")
                documentos_ja_conferidos.add(doc_norm)
                gerar_log(pasta_backup, arquivo, doc_input, "JÁ CONFERIDO")
                continue

            # Busca instantânea no índice
            if doc_norm in indice_docs:
                celula = indice_docs[doc_norm]
                valor = str(celula.value)

                # Se já marcado, avisa
                if f"[✅ {doc_input}]" in valor or "[✅" in valor and doc_norm in documentos_encontrados:
                    print(f"⚠️ Documento {doc_input} já estava marcado na planilha.")
                    documentos_ja_conferidos.add(doc_norm)
                    gerar_log(pasta_backup, arquivo, doc_input, "JÁ MARCADO")
                    continue

                # Marca o número exato dentro da célula
                novos_docs = []
                for d in [d.strip() for d in valor.split(",") if d.strip()]:
                    if normalizar(d) == doc_norm:
                        novos_docs.append(f"[✅ {d}]")
                    else:
                        novos_docs.append(d)

                celula.value = ", ".join(novos_docs)
                documentos_encontrados.add(doc_norm)
                print(f"✅ Documento {doc_input} encontrado e marcado!")
                gerar_log(pasta_backup, arquivo, doc_input, "ENCONTRADO")
            else:
                print(f"❌ Documento {doc_input} não encontrado.")
                documentos_nao_encontrados.add(doc_norm)
                gerar_log(pasta_backup, arquivo, doc_input, "NÃO ENCONTRADO")

            wb.save(arquivo.parent / f"{arquivo.stem}_conferido.xlsx")

        # ==========================================================
        # 📋 RELATÓRIO FINAL
        # ==========================================================
        print("\n📋 RELATÓRIO FINAL:")
        print(f"✔️ Encontrados: {len(documentos_encontrados)}")
        print(f"⚠️ Já conferidos: {len(documentos_ja_conferidos)}")
        print(f"❌ Não encontrados: {len(documentos_nao_encontrados)}")

        # Conta quantos documentos ainda não foram marcados na planilha
        nao_marcados = 0
        for doc, celula in indice_docs.items():
            if "[✅" not in str(celula.value):
                nao_marcados += 1

        print(f"📄 Documentos não marcados na planilha: {nao_marcados}")

        print(f"\n💾 Planilha salva como: {arquivo.stem}_conferido.xlsx")
        print("📝 Log atualizado com os resultados.")

    except Exception as e:
        print(f"\n❌ Erro inesperado: {e}")
        sys.exit(1)


# ==========================================================
# 🚀 Execução Principal
# ==========================================================
if __name__ == "__main__":
    Tk().withdraw()

    print("📁 Selecione o arquivo Excel para conferência:")
    caminho_arquivo = filedialog.askopenfilename(
        title="Selecione a planilha Excel",
        filetypes=[("Arquivos Excel", "*.xlsx *.xlsm *.xltx *.xltm")]
    )

    if not caminho_arquivo:
        print("❌ Nenhum arquivo selecionado. Encerrando.")
        sys.exit()

    print("\n📂 Agora selecione a pasta onde deseja salvar o backup e o log:")
    pasta_backup = filedialog.askdirectory(title="Selecione a pasta de backup e log")

    if not pasta_backup:
        print("❌ Nenhuma pasta selecionada. Encerrando.")
        sys.exit()

    conferir_documentos_interativo(caminho_arquivo, Path(pasta_backup))
